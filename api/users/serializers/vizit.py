from datetime import datetime

import openpyxl
import pytz
from rest_framework import serializers
from rest_framework.serializers import ModelSerializer
from os.path import join as join_path

from DivineBase.settings import BASE_DIR
from apps.users.models import User, District, City, Hospital, Doctor
from apps.vizit.models import Vizit, VizitExcel

EXCEL_PATH = join_path(BASE_DIR, 'media', 'excel_utils')


def check_excel(file_path, excel_path):
    try:
        with open(f'{join_path(file_path, excel_path)}', 'r') as file:
            return "bosingiz"
    except BaseException:
        return "file yoq"


def create_doctor_all_excel(created_at: object, created_by: object, phone_number: object, village: object, city: object,
                            lpu: object, doctor_name: object, category, d_type, doctor_phone: object,
                            comment: object) -> object:
    village = str(village).replace(" ", '_')
    excel_path = f"obshi_vizit_excel_{village}.xlsx"
    check_excel_params = check_excel(EXCEL_PATH, excel_path)
    if check_excel_params.__eq__("file yoq"):
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Vaqti", 'Kimdan', 'Telefon Nomer', "Viloyat", 'Shaxar', 'LPU', 'Doktor ismi', 'Mutaxasisligi',
                   "Kategoriyasi",
                   "Doktor telefoni",
                   'Izoh']
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        sheet1["E1"] = columns[4]
        sheet1["F1"] = columns[5]
        sheet1["G1"] = columns[6]
        sheet1['H1'] = columns[7]
        sheet1['I1'] = columns[8]
        sheet1['J1'] = columns[9]
        sheet1['K1'] = columns[10]
        book1.save(f'{join_path(EXCEL_PATH, excel_path)}')
        book = openpyxl.load_workbook(f'{join_path(EXCEL_PATH, excel_path)}')
        sheet = book.active
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 30
        sheet.column_dimensions['J'].width = 30
        sheet.column_dimensions['K'].width = 60
        row = (
            created_at, created_by, phone_number, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment)
        sheet.append(row)
        book.save(f'{join_path(EXCEL_PATH, excel_path)}')
    else:
        book = openpyxl.load_workbook(f'{join_path(EXCEL_PATH, excel_path)}')
        sheet = book.active
        row = (
            created_at, created_by, phone_number, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment)
        sheet.append(row)
        book.save(f'{join_path(EXCEL_PATH, excel_path)}')


class DistrictVizitSerializer(ModelSerializer):
    class Meta:
        model = District
        fields = '__all__'


class UserVizitSerializer(ModelSerializer):
    district = DistrictVizitSerializer()

    class Meta:
        model = User
        fields = '__all__'


class DoctorVizitSerializer(ModelSerializer):
    class Meta:
        model = Doctor
        fields = '__all__'


class HospitalVizitSerializer(ModelSerializer):
    class Meta:
        model = Hospital
        fields = '__all__'


class CityVizitSerializer(ModelSerializer):
    class Meta:
        model = City
        fields = '__all__'


class GetVizitSerializer(ModelSerializer):
    lpu = HospitalVizitSerializer()
    user = UserVizitSerializer()
    city = CityVizitSerializer()
    doctor = DoctorVizitSerializer()

    class Meta:
        model = Vizit
        fields = ['id', 'lpu', 'user', 'city', 'doctor', 'comment', 'created_date']


# class District
class VizitExcelSerializer(ModelSerializer):
    district = DistrictVizitSerializer()
    path = serializers.FileField()

    class Meta:
        model = VizitExcel
        fields = ['id', 'path', 'district']


class VizitSerializer(ModelSerializer):
    user = serializers.PrimaryKeyRelatedField(queryset=User.objects.all())
    city = serializers.PrimaryKeyRelatedField(queryset=City.objects.all())
    lpu = serializers.PrimaryKeyRelatedField(queryset=Hospital.objects.all())
    doctor = serializers.PrimaryKeyRelatedField(queryset=Doctor.objects.all())
    comment = serializers.CharField(max_length=1000)
    district = serializers.PrimaryKeyRelatedField(queryset=District.objects.all(), required=False)

    class Meta:
        model = Vizit
        fields = [
            'user',
            'city',
            'lpu',
            'doctor',
            'comment',
            'district'
        ]

    def validate(self, attrs):
        user = attrs.get('user')
        city = attrs.get('city')
        lpu = attrs.get('lpu')
        doctor = attrs.get('doctor')
        created_at = str(pytz.timezone('Asia/Tashkent').localize(datetime.now()))[:19]
        created_by = user.first_name
        if user.last_name:
            created_by = f"{user.first_name} {user.last_name}"
        village = user.district.name
        city = city.name
        lpu = lpu.name
        doctor_name = doctor.name
        category = doctor.category_doctor
        d_type = doctor.type_doctor
        doctor_phone = doctor.phone_number
        phone_number = user.phone_number
        create_doctor_all_excel(created_at=created_at, created_by=created_by, phone_number=phone_number,
                                village=village, city=city, lpu=lpu, doctor_name=doctor_name, category=category,
                                d_type=d_type, doctor_phone=doctor_phone, comment=attrs.get('comment'))
        user = User.objects.get(id=user.id)
        district = user.district
        attrs['district'] = district
        Vizit.objects.create(**attrs)
        return attrs
