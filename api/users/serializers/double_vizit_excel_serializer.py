from datetime import datetime

import pytz
from rest_framework import serializers
import openpyxl

from DivineBase.settings import BASE_DIR
from api.users.serializers.vizit import check_excel
from apps.double_vizit.models import DoubleVizitExcel
from os.path import join as join_path

from apps.users.models import User, CheckMp

EXCEL_PATH = join_path(BASE_DIR, 'media', 'double_vizit')


def create_mp_doctor_or_pharmacy_all_excel(created_at,
                                           created_by,
                                           village,
                                           city,
                                           lpu,
                                           doctor_name,
                                           category,
                                           d_type,
                                           doctor_phone,
                                           comment,
                                           pharmacy_name,
                                           pharmacy_address,
                                           mp=None,
                                           preparation=1,
                                           communication=1,
                                           the_need=1,
                                           presentation=1,
                                           protest=1,
                                           agreement=1,
                                           analysis=1,

                                           ) -> None:
    if preparation != "Ma'lumot yo'q":
        urtacha = (int(presentation[0]) + int(preparation[0]) + int(communication[0]) + int(the_need[0]) + int(
            protest[0]) + int(
            agreement[0]) + int(analysis[0])) / 7
        presentation = presentation[0]
        preparation = preparation[0]
        communication = communication[0]
        the_need = the_need[0]
        protest = protest[0]
        agreement = agreement[0]
        analysis = analysis[0]
    else:
        urtacha = "Ma'lumot yo'q"
    excel_path = "pharmacy_or_vizit.xlsx"
    year = created_at[:11]
    hour = created_at[11:19]
    check_excel_params = check_excel(EXCEL_PATH, excel_path)
    if check_excel_params.__eq__("file yoq"):
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Vaqti(Yil-Oy-Kun)", "Vaqti(Soat-Minut)", 'Kimdan', "Viloyat",
                   'Shaxar', 'LPU', 'Doktor ismi',
                   'Mutaxasisligi', "Kategoriyasi",
                   "Doktor telefoni", 'Izoh', "Dorixona Nomi",
                   "Dorixona Manzili", "MP",
                   "Tayyorgarlik", 'Muloqot', 'Extiyoj',
                   'Taqdimot', "E'tiroz",
                   'Kelishuv', 'Taxlil', "O'rtacha"]
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        sheet1["E1"] = columns[4]
        sheet1["F1"] = columns[5]
        sheet1["G1"] = columns[6]
        sheet1['H1'] = columns[7]
        sheet1['I1'] = columns[8]
        sheet1['J1'] = columns[9]
        sheet1['K1'] = columns[10]
        sheet1['L1'] = columns[11]
        sheet1['M1'] = columns[12]
        sheet1['N1'] = columns[13]
        sheet1['O1'] = columns[14]
        sheet1['P1'] = columns[15]
        sheet1['Q1'] = columns[16]
        sheet1['R1'] = columns[17]
        sheet1['S1'] = columns[18]
        sheet1['T1'] = columns[19]
        sheet1['U1'] = columns[20]
        sheet1['V1'] = columns[21]
        book1.save(f'{join_path(EXCEL_PATH, excel_path)}')
        book = openpyxl.load_workbook(f'{join_path(EXCEL_PATH, excel_path)}')
        sheet = book.active
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 30
        sheet.column_dimensions['J'].width = 30
        sheet.column_dimensions['K'].width = 60
        sheet.column_dimensions['L'].width = 30
        sheet.column_dimensions['M'].width = 30
        sheet.column_dimensions['N'].width = 30
        sheet.column_dimensions['O'].width = 30
        sheet.column_dimensions['P'].width = 30
        sheet.column_dimensions['Q'].width = 30
        sheet.column_dimensions['R'].width = 30
        sheet.column_dimensions['S'].width = 30
        sheet.column_dimensions['T'].width = 30
        sheet.column_dimensions['U'].width = 30
        sheet.column_dimensions['V'].width = 30
        row = (
            year, hour, created_by, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment, pharmacy_name, pharmacy_address, mp, preparation, communication, the_need, presentation, protest,
            agreement, analysis, urtacha)
        sheet.append(row)
        book.save(f'{join_path(EXCEL_PATH, excel_path)}')
    else:
        book = openpyxl.load_workbook(f'{join_path(EXCEL_PATH, excel_path)}')
        sheet = book.active
        row = (
            year, hour, created_by, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment, pharmacy_name, pharmacy_address, mp, preparation, communication, the_need, presentation, protest,
            agreement, analysis, urtacha)
        sheet.append(row)
        book.save(f'{join_path(EXCEL_PATH, excel_path)}')


class DoubleVizitSerializer(serializers.ModelSerializer):
    class Meta:
        fields = "__all__"
        model = DoubleVizitExcel
